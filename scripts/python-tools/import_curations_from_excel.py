#!/usr/bin/env python3
"""
File: import_curations_from_excel.py
Purpose: Import curation documents directly into MongoDB from Excel files.
Dependencies: pymongo, openpyxl

Main Responsibilities:
- Read all .xlsx files from a folder
- Extract category concepts from spreadsheet rows
- Build normalized curation documents for API V3 schema
- Support dry-run preview and apply mode (write to MongoDB)
"""

import argparse
import hashlib
import os
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from pymongo import MongoClient


def parse_args() -> argparse.Namespace:
    default_input = Path(__file__).resolve().parents[2] / "data" / "aiconciergerevisorestaurantessopaulomaro2023"
    parser = argparse.ArgumentParser(description="Import curations from Excel files into MongoDB")
    parser.add_argument("--input-dir", type=Path, default=default_input, help="Folder containing .xlsx files")
    parser.add_argument("--apply", action="store_true", help="Write to MongoDB (default is dry-run)")
    parser.add_argument("--default-curator-id", default="curator-import-excel", help="Curator ID for imported curations")
    parser.add_argument("--default-curator-name", default="Excel Import Script", help="Curator name for imported curations")
    return parser.parse_args()


def load_env_values() -> Dict[str, str]:
    env_values: Dict[str, str] = {}
    root = Path(__file__).resolve().parents[2]
    env_path = root / "concierge-api-v3" / ".env"

    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            text = line.strip()
            if not text or text.startswith("#") or "=" not in text:
                continue
            key, value = text.split("=", 1)
            env_values[key.strip()] = value.strip().strip('"').strip("'")

    return env_values


def get_mongo_connection() -> Tuple[MongoClient, str]:
    env_values = load_env_values()
    mongo_uri = env_values.get("MONGODB_URL") or env_values.get("MONGODB_URI") or env_values.get("MONGO_URI") or os.getenv("MONGODB_URL") or os.getenv("MONGODB_URI") or os.getenv("MONGO_URI")
    db_name = env_values.get("MONGODB_DB_NAME") or env_values.get("MONGO_DB_NAME") or env_values.get("DB_NAME") or os.getenv("MONGODB_DB_NAME") or os.getenv("MONGO_DB_NAME") or os.getenv("DB_NAME") or "concierge_collector"

    if not mongo_uri:
        raise RuntimeError("Mongo URI not found. Expected MONGODB_URL/MONGODB_URI/MONGO_URI in concierge-api-v3/.env")

    return MongoClient(mongo_uri), db_name


def normalize_value(value: Any) -> List[str]:
    if value is None:
        return []

    raw = str(value).strip()
    if not raw:
        return []

    separators = [";", "|", "\n"]
    for separator in separators:
        raw = raw.replace(separator, ",")

    parts = [part.strip() for part in raw.split(",")]
    return [part for part in parts if part]


def build_categories(sheet) -> Dict[str, List[str]]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}

    headers = [str(item).strip() if item is not None else "" for item in header_row]
    category_values: Dict[str, OrderedDict[str, None]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        for idx, header in enumerate(headers):
            if not header:
                continue
            if idx >= len(row):
                continue

            values = normalize_value(row[idx])
            if not values:
                continue

            if header not in category_values:
                category_values[header] = OrderedDict()

            for entry in values:
                lower = entry.lower()
                if lower not in category_values[header]:
                    category_values[header][lower] = entry

    return {
        category: list(unique_map.values())
        for category, unique_map in category_values.items()
        if unique_map
    }


def build_curation_document(
    excel_path: Path,
    sheet_name: str,
    categories: Dict[str, List[str]],
    curator_id: str,
    curator_name: str,
) -> Dict[str, Any]:
    base_id = f"excel::{excel_path.name}::{sheet_name}".lower()
    digest = hashlib.md5(base_id.encode("utf-8")).hexdigest()[:12]
    curation_id = f"curation-{digest}"

    return {
        "_id": curation_id,
        "curation_id": curation_id,
        "restaurant_name": sheet_name.strip() if sheet_name else excel_path.stem,
        "status": "draft",
        "notes": {
            "public": f"Curated concepts for {sheet_name.strip() if sheet_name else excel_path.stem}",
            "private": f"Imported from Excel {excel_path.name} at {datetime.now(timezone.utc).isoformat()}"
        },
        "categories": categories,
        "sources": {
            "import": [
                {
                    "source": excel_path.name,
                    "sheet": sheet_name,
                    "format": "xlsx"
                }
            ]
        },
        "items": [],
        "curator_id": curator_id,
        "curator": {
            "id": curator_id,
            "name": curator_name,
            "email": None
        }
    }


def extract_from_excel(excel_path: Path, curator_id: str, curator_name: str) -> List[Dict[str, Any]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    curations: List[Dict[str, Any]] = []

    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            categories = build_categories(sheet)
            curation = build_curation_document(excel_path, sheet_name, categories, curator_id, curator_name)
            curations.append(curation)
    finally:
        workbook.close()

    return curations


def save_curation(collection, curation: Dict[str, Any]) -> str:
    now = datetime.now(timezone.utc)

    update_result = collection.update_one(
        {"_id": curation["_id"]},
        {
            "$set": {
                "curation_id": curation["curation_id"],
                "restaurant_name": curation["restaurant_name"],
                "status": curation["status"],
                "notes": curation["notes"],
                "categories": curation["categories"],
                "sources": curation["sources"],
                "items": curation["items"],
                "curator_id": curation["curator_id"],
                "curator": curation["curator"],
                "updatedAt": now,
            },
            "$setOnInsert": {
                "createdAt": now,
                "version": 1,
            },
        },
        upsert=True,
    )

    return "created" if update_result.upserted_id else "updated"


def main() -> int:
    args = parse_args()
    input_dir: Path = args.input_dir

    if not input_dir.exists() or not input_dir.is_dir():
        print(f"ERROR: input folder not found: {input_dir}")
        return 1

    excel_files = sorted(input_dir.glob("*.xlsx"))
    if not excel_files:
        print(f"No .xlsx files found in {input_dir}")
        return 0

    print(f"Found {len(excel_files)} Excel file(s)")
    print(f"Mode: {'APPLY' if args.apply else 'DRY-RUN'}")

    client = None
    collection = None
    if args.apply:
        client, db_name = get_mongo_connection()
        collection = client[db_name].curations

    created = 0
    updated = 0
    failed = 0
    total_sheets = 0

    for index, excel_path in enumerate(excel_files, start=1):
        try:
            curations = extract_from_excel(excel_path, args.default_curator_id, args.default_curator_name)
            total_sheets += len(curations)

            for sheet_index, curation in enumerate(curations, start=1):
                concepts_count = sum(len(values) for values in curation["categories"].values())

                if not args.apply:
                    print(
                        f"[{index}/{len(excel_files)} | {sheet_index}/{len(curations)}] "
                        f"DRY-RUN {curation['curation_id']} | name={curation['restaurant_name']} "
                        f"| categories={len(curation['categories'])} | concepts={concepts_count}"
                    )
                    continue

                outcome = save_curation(collection, curation)
                if outcome == "created":
                    created += 1
                else:
                    updated += 1

                print(
                    f"[{index}/{len(excel_files)} | {sheet_index}/{len(curations)}] "
                    f"{outcome.upper()} {curation['curation_id']} | name={curation['restaurant_name']} "
                    f"| categories={len(curation['categories'])} | concepts={concepts_count}"
                )

        except Exception as error:
            failed += 1
            print(f"[{index}/{len(excel_files)}] FAILED {excel_path.name} -> {error}")

    if client:
        client.close()

    print("\n" + "=" * 72)
    print("Summary")
    print("=" * 72)
    print(f"Total files: {len(excel_files)}")
    print(f"Total sheets (restaurants): {total_sheets}")
    print(f"Created: {created}")
    print(f"Updated: {updated}")
    print(f"Failed: {failed}")

    if not args.apply:
        print("\nDry-run only. Use --apply to write into MongoDB.")

    return 0 if failed == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
